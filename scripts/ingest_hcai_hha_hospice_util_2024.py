# ==============================================================================
# file_id: SOM-SCR-0022-v0.1.0
# name: ingest_hcai_hha_hospice_util_2024.py
# description: Ingest HCAI 2024 Home Health & Hospice utilization XLSX into a clean CSV keyed by FAC_NO (HCAI facility id)
# project_id: HIPPOCRATIC
# category: script
# tags: [hcai, hospice, home-health, utilization, medi-cal, revenue, ingest]
# created: 2026-01-16
# modified: 2026-01-16
# version: 0.1.0
# agent_id: AGENT-CURSOR-OPENAI
# execution: .venv\\Scripts\\python scripts/ingest_hcai_hha_hospice_util_2024.py
# ==============================================================================

from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
IN_XLSX = ROOT / "data" / "source" / "hcai" / "hha_hospice_util" / "2024-complete-data-set-home-health-and-hospice-october-2025.xlsx"
OUT_CSV = ROOT / "data" / "enrichment" / "state" / "CA" / "hcai_hhah_util_2024.csv"

SHEET = "Page 1-11"

# Keep this tight for the first iteration: enough to join + show “Medi-Cal vs Medicare revenue/volume” signals.
KEEP_COLS = [
    "FAC_NO",
    "FAC_NAME",
    "FAC_STR_ADDR",
    "FAC_CITY",
    "FAC_ZIP",
    "FAC_PHONE",
    "LICENSE_NO",
    "LICENSE_STATUS",
    "LIC_CAT",
    "HHAH_LICEE_TOC",  # "Home Health Agency Only" vs "Hospice Only" (etc.)
    "HHAH_MEDI_CAL_VISITS",
    "HHAH_MEDICARE_VISITS",
    "HOSPICE_PATS_PAID_BY_MEDI_CAL",
    "HOSPICE_PATS_PAID_BY_MEDICARE",
    "HOSPICE_MEDI_CAL_REVENUE",
    "HOSPICE_MEDICARE_REVENUE",
    "HOSPICE_TOT_OPER_REVENUE",
    "HOSPICE_NET_INCOME",
]


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass
    return str(v).strip()


def main() -> None:
    if not IN_XLSX.exists():
        raise SystemExit(f"Input not found: {IN_XLSX}")

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(IN_XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET]

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col_to_idx = {str(name): i for i, name in enumerate(header)}

    missing = [c for c in KEEP_COLS if c not in col_to_idx]
    if missing:
        raise SystemExit(f"Missing expected columns: {missing}")

    exported_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    out_fields = ["year", "exported_at"] + KEEP_COLS
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=out_fields)
        w.writeheader()

        for row in ws.iter_rows(min_row=6, values_only=True):  # row 6 is first data row in this sheet
            fac_no = row[col_to_idx["FAC_NO"]]
            if fac_no is None:
                continue
            record = {"year": "2024", "exported_at": exported_at}
            for c in KEEP_COLS:
                record[c] = _to_str(row[col_to_idx[c]])
            w.writerow(record)

    print(str(OUT_CSV))


if __name__ == "__main__":
    main()




